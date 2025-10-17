import asyncio
import json
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Optional

# Importamos la nueva librería
from bitunix import BitunixClient

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scipy.signal import argrelextrema
from telegram import Bot
from telegram.error import TelegramError

# --- Configuración del Logging ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('smc_trading_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class Position:
    """Estructura para almacenar el estado de una posición SMC."""
    direction: str
    size: float
    entry_price: float
    entry_time: datetime
    entry_idx: int
    stop_loss: float
    take_profit: float
    liquidation_price: float
    margin_used: float

class SmartMoneyLiveBot:
    """Bot de trading en vivo que implementa la estrategia SMC usando BitunixClient."""

    def __init__(
        self,
        api_key: str,
        api_secret: str,
        telegram_token: str,
        telegram_chat_id: str,
        symbol: str = 'BTCUSDT',
        initial_balance: float = 1000.0,
    ) -> None:
        self.symbol = symbol
        self.timeframe = '5m'
        self.candle_limit = 300
        self.refresh_seconds = 15

        # Parámetros de la Estrategia SMC
        self.structure_lookback = 20
        self.risk_reward_ratio = 2.5
        self.leverage = 20
        self.risk_per_trade_pct = 0.02

        # Telegram
        self.telegram_bot = Bot(token=telegram_token)
        self.telegram_chat_id = telegram_chat_id
        
        # Estado y Reportes
        self.initial_balance = initial_balance
        self.balance = initial_balance
        self.trades: list[Dict] = []
        self.excel_filename = f"live_report_SMC_{symbol.replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self._initialize_excel()

        # Conexión con BitunixClient
        logger.info("Conectando a Bitunix...")
        self.client = BitunixClient(api_key=api_key, api_secret=api_secret)

        self.df: Optional[pd.DataFrame] = None
        self.position: Optional[Position] = None
        self.is_running: bool = False
        self.last_signal_time: Optional[pd.Timestamp] = None

    # --- Métodos de Reporte en Excel (sin cambios) ---
    def _initialize_excel(self) -> None:
        wb = Workbook()
        ws_trades = wb.active
        ws_trades.title = "Trades"
        headers = [
            'Fecha Entrada', 'Hora Entrada', 'Fecha Salida', 'Hora Salida', 'Dirección',
            'Precio Entrada', 'Precio Salida', 'Stop Loss', 'Take Profit', 'Liquidación',
            'Tamaño Posición', 'Margen Usado', 'P/L USD', 'Razón Salida', 'Balance'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        wb.save(self.excel_filename)
        logger.info(f"📊 Archivo Excel creado: {self.excel_filename}")

    def _save_trade_to_excel(self, trade: Dict) -> None:
        try:
            wb = load_workbook(self.excel_filename)
            ws = wb["Trades"]
            
            entry_time = trade['entry_time']
            exit_time = trade['exit_time']
            
            row_data = [
                entry_time.strftime('%Y-%m-%d'), entry_time.strftime('%H:%M:%S'),
                exit_time.strftime('%Y-%m-%d'), exit_time.strftime('%H:%M:%S'),
                trade['direction'], trade['entry_price'], trade['exit_price'],
                trade['stop_loss'], trade['take_profit'], trade['liquidation_price'],
                trade['position_size'], trade['margin_used'], trade['pnl'],
                trade['exit_reason'], self.balance
            ]
            ws.append(row_data)
            
            pnl_cell = ws.cell(row=ws.max_row, column=13)
            if trade['pnl'] > 0:
                pnl_cell.font = Font(color="00B050", bold=True)
            elif trade['pnl'] < 0:
                pnl_cell.font = Font(color="FF0000", bold=True)

            wb.save(self.excel_filename)
            logger.info(f"📊 Trade guardado en Excel.")
        except Exception as e:
            logger.error(f"Error al guardar en Excel: {e}")

    # --- Métodos de Notificación por Telegram (sin cambios) ---
    async def send_telegram_message(self, message: str) -> None:
        try:
            await self.telegram_bot.send_message(chat_id=self.telegram_chat_id, text=message, parse_mode='HTML')
        except TelegramError as e:
            logger.error(f"Error al enviar mensaje de Telegram: {e}")

    async def notify_entry(self, pos: Position) -> None:
        pnl_target = (pos.take_profit - pos.entry_price) * pos.size if pos.direction == 'LONG' else (pos.entry_price - pos.take_profit) * pos.size
        msg = (
            f"✅ <b>POSICIÓN ABIERTA (SMC)</b>\n\n"
            f"📊 Par: {self.symbol}\n"
            f"📈 Dirección: <b>{pos.direction}</b>\n"
            f"💰 Precio entrada: ${pos.entry_price:,.4f}\n"
            f"📏 Tamaño: {pos.size:.4f}\n"
            f"💵 Margen: ${pos.margin_used:,.2f}\n"
            f"🎯 Take Profit: ${pos.take_profit:,.4f} (+${pnl_target:,.2f})\n"
            f"🛑 Stop Loss: ${pos.stop_loss:,.4f}\n"
            f"⚠️ Liquidación: ${pos.liquidation_price:,.4f}"
        )
        await self.send_telegram_message(msg)

    async def notify_exit(self, trade: Dict) -> None:
        emoji = "🟢" if trade['pnl'] > 0 else "🔴"
        msg = (
            f"{emoji} <b>POSICIÓN CERRADA (SMC)</b>\n\n"
            f"📊 Par: {self.symbol}\n"
            f"📈 Dirección: {trade['direction']}\n"
            f"💰 Entrada: ${trade['entry_price']:,.4f} | Salida: ${trade['exit_price']:,.4f}\n"
            f"💵 P/L: <b>${trade['pnl']:,.2f}</b>\n"
            f"📝 Razón: {trade['exit_reason']}\n"
            f"💼 Balance: ${self.balance:,.2f}"
        )
        await self.send_telegram_message(msg)

    # --- Lógica de la Estrategia SMC ---
    async def update_market_data(self) -> bool:
        """Actualiza los datos de mercado y calcula los patrones SMC usando BitunixClient."""
        try:
            # --- CORRECCIÓN: Usar la función correcta para klines de futuros ---
            ohlcv = self.client.get_kline_data(symbol=self.symbol, interval=self.timeframe)

            
            df = pd.DataFrame(ohlcv, columns=[
                'timestamp', 'open', 'high', 'low', 'close', 'volume', 'close_time', 
                'quote_asset_volume', 'number_of_trades', 'taker_buy_base_asset_volume', 
                'taker_buy_quote_asset_volume', 'ignore'
            ])
            
            df = df[['timestamp', 'open', 'high', 'low', 'close', 'volume']]
            df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
            df.set_index('timestamp', inplace=True)
            df = df.astype(float)

            n = self.structure_lookback
            df['min'] = df.iloc[argrelextrema(df.low.values, np.less_equal, order=n)[0]]['low']
            df['max'] = df.iloc[argrelextrema(df.high.values, np.greater_equal, order=n)[0]]['high']

            df['fvg_bull_high'], df['fvg_bull_low'] = np.nan, np.nan
            df['fvg_bear_high'], df['fvg_bear_low'] = np.nan, np.nan

            for i in range(2, len(df)):
                if df['low'].iloc[i] > df['high'].iloc[i-2]:
                    df.loc[df.index[i-1], 'fvg_bull_low'] = df['high'].iloc[i-2]
                    df.loc[df.index[i-1], 'fvg_bull_high'] = df['low'].iloc[i]
                if df['high'].iloc[i] < df['low'].iloc[i-2]:
                    df.loc[df.index[i-1], 'fvg_bear_low'] = df['high'].iloc[i]
                    df.loc[df.index[i-1], 'fvg_bear_high'] = df['low'].iloc[i-2]
            
            self.df = df
            return True
        except Exception as e:
            logger.error(f"Error actualizando datos de mercado con Bitunix: {e}")
            return False

    def check_long_setup(self, i: int) -> bool:
        recent_lows = self.df['min'].iloc[i-50:i].dropna()
        if len(recent_lows) < 2 or recent_lows.iloc[-1] >= recent_lows.iloc[-2]: return False
        
        sweep_idx = self.df.index.get_loc(recent_lows.index[-1])
        if i - sweep_idx > 12: return False
        
        fvg_window = self.df.iloc[sweep_idx:i]
        bullish_fvgs = fvg_window[['fvg_bull_low', 'fvg_bull_high']].dropna()
        if bullish_fvgs.empty: return False
        
        last_fvg = bullish_fvgs.iloc[-1]
        if self.df['low'].iloc[i] <= last_fvg['fvg_bull_high']:
            entry_price = last_fvg['fvg_bull_high']
            liquidity_level = recent_lows.iloc[-1]
            asyncio.create_task(self.open_position(i, 'LONG', entry_price, liquidity_level))
            return True
        return False

    def check_short_setup(self, i: int) -> bool:
        recent_highs = self.df['max'].iloc[i-50:i].dropna()
        if len(recent_highs) < 2 or recent_highs.iloc[-1] <= recent_highs.iloc[-2]: return False

        sweep_idx = self.df.index.get_loc(recent_highs.index[-1])
        if i - sweep_idx > 12: return False

        fvg_window = self.df.iloc[sweep_idx:i]
        bearish_fvgs = fvg_window[['fvg_bear_low', 'fvg_bear_high']].dropna()
        if bearish_fvgs.empty: return False
            
        last_fvg = bearish_fvgs.iloc[-1]
        if self.df['high'].iloc[i] >= last_fvg['fvg_bear_low']:
            entry_price = last_fvg['fvg_bear_low']
            liquidity_level = recent_highs.iloc[-1]
            asyncio.create_task(self.open_position(i, 'SHORT', entry_price, liquidity_level))
            return True
        return False

    # --- Gestión de Órdenes y Posiciones (sin cambios) ---
    async def open_position(self, entry_idx: int, direction: str, entry_price: float, liquidity_level: float):
        if direction == 'LONG':
            stop_loss_price = liquidity_level * 0.999
            risk_per_unit = entry_price - stop_loss_price
            take_profit_price = entry_price + (risk_per_unit * self.risk_reward_ratio)
        else:
            stop_loss_price = liquidity_level * 1.001
            risk_per_unit = stop_loss_price - entry_price
            take_profit_price = entry_price - (risk_per_unit * self.risk_reward_ratio)

        if risk_per_unit <= 0: return

        capital_to_risk = self.balance * self.risk_per_trade_pct
        position_size = capital_to_risk / risk_per_unit
        margin_used = (position_size * entry_price) / self.leverage
        
        liquidation_pct = 1 / self.leverage * 0.9
        liquidation_price = entry_price * (1 - liquidation_pct) if direction == 'LONG' else entry_price * (1 + liquidation_pct)
        
        self.position = Position(
            direction=direction, size=position_size, entry_price=entry_price,
            entry_time=self.df.index[entry_idx].to_pydatetime(), entry_idx=entry_idx,
            stop_loss=stop_loss_price, take_profit=take_profit_price,
            liquidation_price=liquidation_price, margin_used=margin_used
        )
        self.last_signal_time = self.df.index[entry_idx]
        logger.info(f"📢 Señal {direction}: tamaño {position_size:.4f} @ {entry_price:.4f}")
        await self.notify_entry(self.position)

    async def close_position(self, exit_price: float, reason: str):
        if not self.position: return

        pos = self.position
        pnl = (exit_price - pos.entry_price) * pos.size if pos.direction == 'LONG' else (pos.entry_price - exit_price) * pos.size
        
        self.balance += pnl
        
        trade = {
            'entry_time': pos.entry_time, 'exit_time': datetime.now(),
            'direction': pos.direction, 'entry_price': pos.entry_price,
            'exit_price': exit_price, 'stop_loss': pos.stop_loss,
            'take_profit': pos.take_profit, 'liquidation_price': pos.liquidation_price,
            'position_size': pos.size, 'margin_used': pos.margin_used, 'pnl': pnl,
            'exit_reason': reason
        }
        self.trades.append(trade)
        
        logger.info(f"ℹ️ Posición cerrada ({reason}). P/L: ${pnl:.2f}. Balance: ${self.balance:.2f}")
        self._save_trade_to_excel(trade)
        await self.notify_exit(trade)
        
        self.position = None

    async def manage_position(self, idx: int):
        if not self.position: return

        candle = self.df.iloc[idx]
        pos = self.position

        if pos.direction == 'LONG':
            if candle['low'] <= pos.stop_loss: await self.close_position(pos.stop_loss, 'Stop Loss')
            elif candle['high'] >= pos.take_profit: await self.close_position(pos.take_profit, 'Take Profit')
        else: # SHORT
            if candle['high'] >= pos.stop_loss: await self.close_position(pos.stop_loss, 'Stop Loss')
            elif candle['low'] <= pos.take_profit: await self.close_position(pos.take_profit, 'Take Profit')
        
        if self.position and (idx - pos.entry_idx) >= 24: # max_candles_in_trade
            await self.close_position(candle['close'], 'Time Limit')

    # --- Bucle Principal del Bot (sin cambios) ---
    async def run(self):
        await self.send_telegram_message(f"🚀 <b>Bot SMC Iniciado</b>\nOperando en {self.symbol}...")
        self.is_running = True
        
        while self.is_running:
            try:
                if not await self.update_market_data():
                    await asyncio.sleep(self.refresh_seconds)
                    continue

                idx = len(self.df) - 2 

                if self.position:
                    await self.manage_position(idx)
                else:
                    if self.last_signal_time and self.df.index[idx] == self.last_signal_time:
                        await asyncio.sleep(self.refresh_seconds)
                        continue

                    if not self.check_long_setup(idx):
                        self.check_short_setup(idx)
                
                await asyncio.sleep(self.refresh_seconds)
            except Exception as e:
                logger.error(f"Error en el bucle principal: {e}")
                await self.send_telegram_message(f"🚨 <b>Error Crítico en el Bot</b>\n\n{e}\n\nReintentando en 60 segundos.")
                await asyncio.sleep(60)

    def stop(self):
        self.is_running = False

async def main():
    script_dir = os.path.dirname(os.path.realpath(__file__))
    config_path = os.path.join(script_dir, 'config.json')

    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        logger.error(f"No se encontró 'config.json' en la ruta: {config_path}. Por favor, crea uno.")
        return

    bot = SmartMoneyLiveBot(
        api_key=config['api_key'],
        api_secret=config['api_secret'],
        telegram_token=config['telegram_token'],
        telegram_chat_id=config['telegram_chat_id'],
        symbol=config.get('symbol', 'BTCUSDT'),
        initial_balance=float(config.get('initial_balance', 1000.0))
    )
    
    try:
        await bot.run()
    except KeyboardInterrupt:
        logger.info("Deteniendo bot...")
        await self.send_telegram_message("🛑 <b>Bot Detenido Manualmente</b>")
        bot.stop()

if __name__ == "__main__":
    asyncio.run(main())