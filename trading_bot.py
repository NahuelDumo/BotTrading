"""Bot de trading en vivo para Binance Futures replicando la estrategia híbrida.

Este módulo implementa la lógica EXACTA utilizada en el backtest `HybridTradingBacktest`
para operar en tiempo real. Usa ccxt contra Binance USDT-M Futures, maneja apalancamiento
15x, límites de tiempo diferenciales y salidas por stop, take profit o liquidación.
"""

import asyncio
import json
import logging
import os
from dataclasses import dataclass
from datetime import datetime, time
from typing import Dict, Optional

import ccxt
import numpy as np
import pandas as pd
from telegram import Bot
from telegram.error import TelegramError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# Configuración de logging en español
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('trading_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class Position:
    """Estructura para almacenar el estado de una posición abierta."""

    direction: str
    size: float
    entry_price: float
    entry_time: datetime
    entry_idx: int
    stop_loss: float
    take_profit: float
    liquidation_price: float
    is_first_trade: bool
    margin_used: float
    entry_type: str  # 'mean_reversion' o 'momentum'
    trailing_stop: Optional[float] = None  # Para posiciones de momentum


class HybridFuturesTradingBot:
    """Bot de trading en vivo que replica `HybridTradingBacktest`."""

    def __init__(
        self,
        api_key: str,
        api_secret: str,
        telegram_token: str,
        telegram_chat_id: str,
        symbol: str = 'HYPE/USDT',
        leverage: int = 15,
        risk_per_trade: float = 0.10,
        initial_balance: float = 1_000.0,
        timeframe: str = '5m',
        refresh_seconds: int = 60
    ) -> None:
        self.symbol = symbol
        self.timeframe = timeframe
        self.candle_limit = 600
        self.refresh_seconds = refresh_seconds

        # Parámetros del backtest
        self.leverage = leverage
        self.stop_loss_percent_long = 0.012
        self.stop_loss_percent_short = 0.015
        self.take_profit_percent_long = 0.03
        self.take_profit_percent_short = 0.022
        self.max_candles_long = 60
        self.max_candles_short = 45
        self.risk_per_trade = risk_per_trade

        # Telegram Bot
        self.telegram_token = telegram_token
        self.telegram_chat_id = telegram_chat_id
        self.telegram_bot = Bot(token=telegram_token)
        
        # Estado financiero (simulado según backtest)
        self.initial_balance = initial_balance
        self.balance = initial_balance
        self.trades: list[Dict] = []
        self.equity_curve: list[Dict] = []
        
        # Excel reporting
        self.excel_filename = f"trading_live_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self._initialize_excel()

        # Estado de ejecución
        self.exchange = ccxt.binance({
            'apiKey': api_key,
            'secret': api_secret,
            'enableRateLimit': True,
            'options': {
                'defaultType': 'future',  # USDT-M Futures
                'adjustForTimeDifference': True,
            }
        })
        self.exchange.load_markets()

        try:
            self.exchange.set_leverage(self.leverage, self.symbol)
            logger.info(f"Apalancamiento configurado a {self.leverage}x para {self.symbol}")
        except Exception as exc:
            logger.warning(f"No se pudo fijar apalancamiento automáticamente: {exc}")

        self.df: Optional[pd.DataFrame] = None
        self.position: Optional[Position] = None
        self.is_running: bool = False
        self.filtered_long_signals = 0
        self.filtered_short_signals = 0
        self.daily_first_trade: bool = True
        self.last_trade_date: Optional[datetime.date] = None
        self.last_signal_time: Optional[pd.Timestamp] = None
        
        # Parámetros para momentum trading
        self.momentum_enabled = True
        self.momentum_threshold_pct = 2.0  # Caída/subida de 2% en pocas velas
        self.momentum_candles_window = 3  # Ventana de 3 velas para detectar momentum
        self.momentum_volume_multiplier = 1.5  # Volumen 1.5x mayor que promedio
        self.trailing_stop_activation_pct = 1.5  # Activar trailing después de 1.5% ganancia
        self.trailing_stop_distance_pct = 0.8  # Trailing stop a 0.8% del máximo

    # ----------------------------------------------------------------------------------
    # Excel Reporting
    # ----------------------------------------------------------------------------------
    def _initialize_excel(self) -> None:
        """Inicializar archivo Excel con headers."""
        wb = Workbook()
        
        # Hoja 1: Trades
        ws_trades = wb.active
        ws_trades.title = "Trades"
        
        # Headers para trades
        headers = [
            'Fecha Entrada', 'Hora Entrada', 'Fecha Salida', 'Hora Salida',
            'Dirección', 'Precio Entrada', 'Precio Salida', 
            'Stop Loss', 'Take Profit', 'Liquidación',
            'Tamaño Posición', 'Margen Usado', 'P/L USD', 'P/L %',
            'Razón Salida', 'Duración (velas)', 'Balance'
        ]
        
        # Escribir headers con formato
        for col, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        ws_trades.column_dimensions['A'].width = 12
        ws_trades.column_dimensions['B'].width = 10
        ws_trades.column_dimensions['C'].width = 12
        ws_trades.column_dimensions['D'].width = 10
        ws_trades.column_dimensions['E'].width = 10
        ws_trades.column_dimensions['F'].width = 12
        ws_trades.column_dimensions['G'].width = 12
        ws_trades.column_dimensions['H'].width = 12
        ws_trades.column_dimensions['I'].width = 12
        ws_trades.column_dimensions['J'].width = 12
        ws_trades.column_dimensions['K'].width = 14
        ws_trades.column_dimensions['L'].width = 12
        ws_trades.column_dimensions['M'].width = 10
        ws_trades.column_dimensions['N'].width = 10
        ws_trades.column_dimensions['O'].width = 15
        ws_trades.column_dimensions['P'].width = 15
        ws_trades.column_dimensions['Q'].width = 10
        
        # Hoja 2: Resumen
        ws_summary = wb.create_sheet("Resumen")
        ws_summary['A1'] = "Métrica"
        ws_summary['B1'] = "Valor"
        ws_summary['A1'].font = Font(bold=True)
        ws_summary['B1'].font = Font(bold=True)
        
        summary_data = [
            ['Balance Inicial', self.initial_balance],
            ['Balance Actual', self.initial_balance],
            ['P/L Total', 0],
            ['Retorno %', 0],
            ['Total Trades', 0],
            ['Trades Ganadores', 0],
            ['Trades Perdedores', 0],
            ['Win Rate %', 0],
            ['Última Actualización', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
        
        wb.save(self.excel_filename)
        logger.info(f"📊 Archivo Excel creado: {self.excel_filename}")

    def _save_trade_to_excel(self, trade: Dict) -> None:
        """Guardar trade en Excel."""
        try:
            wb = load_workbook(self.excel_filename)
            ws_trades = wb['Trades']
            
            # Encontrar la siguiente fila vacía
            next_row = ws_trades.max_row + 1
            
            # Preparar datos del trade
            entry_time = trade['entry_time']
            exit_time = trade['exit_time']
            
            row_data = [
                entry_time.strftime('%Y-%m-%d'),
                entry_time.strftime('%H:%M:%S'),
                exit_time.strftime('%Y-%m-%d'),
                exit_time.strftime('%H:%M:%S'),
                trade['direction'],
                round(trade['entry_price'], 4),
                round(trade['exit_price'], 4),
                round(trade['stop_loss'], 4),
                round(trade['take_profit'], 4),
                round(trade['liquidation_price'], 4),
                round(trade['position_size'], 4),
                round(trade.get('margin_used', 0), 2),
                round(trade['pnl'], 2),
                round(trade['pnl_pct'], 2),
                trade['exit_reason'],
                trade['duration_candles'],
                round(self.balance, 2)
            ]
            
            # Escribir datos
            for col, value in enumerate(row_data, 1):
                cell = ws_trades.cell(row=next_row, column=col, value=value)
                
                # Colorear P/L
                if col == 13:  # Columna P/L USD
                    if value > 0:
                        cell.font = Font(color="00B050", bold=True)
                    elif value < 0:
                        cell.font = Font(color="FF0000", bold=True)
            
            # Actualizar resumen
            ws_summary = wb['Resumen']
            winning_trades = sum(1 for t in self.trades if t['pnl'] > 0)
            losing_trades = len(self.trades) - winning_trades
            win_rate = (winning_trades / len(self.trades) * 100) if self.trades else 0
            total_pnl = sum(t['pnl'] for t in self.trades)
            return_pct = ((self.balance - self.initial_balance) / self.initial_balance * 100)
            
            ws_summary['B2'] = self.initial_balance
            ws_summary['B3'] = round(self.balance, 2)
            ws_summary['B4'] = round(total_pnl, 2)
            ws_summary['B5'] = round(return_pct, 2)
            ws_summary['B6'] = len(self.trades)
            ws_summary['B7'] = winning_trades
            ws_summary['B8'] = losing_trades
            ws_summary['B9'] = round(win_rate, 2)
            ws_summary['B10'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            wb.save(self.excel_filename)
            logger.info(f"📊 Trade guardado en Excel: {self.excel_filename}")
            
        except Exception as e:
            logger.error(f"Error al guardar en Excel: {e}")

    # ----------------------------------------------------------------------------------
    # Telegram Notifications
    # ----------------------------------------------------------------------------------
    async def send_telegram_message(self, message: str) -> None:
        """Enviar mensaje a Telegram de forma asíncrona."""
        try:
            await self.telegram_bot.send_message(
                chat_id=self.telegram_chat_id,
                text=message,
                parse_mode='HTML'
            )
        except TelegramError as e:
            logger.error(f"Error al enviar mensaje de Telegram: {e}")
        except Exception as e:
            logger.error(f"Error inesperado en Telegram: {e}")

    async def notify_signal(self, direction: str, price: float, idx: int) -> None:
        """Notificar señal detectada."""
        current = self.df.iloc[idx]
        message = (
            f"🔔 <b>SEÑAL DETECTADA</b>\n\n"
            f"📊 Par: {self.symbol}\n"
            f"📈 Dirección: <b>{direction}</b>\n"
            f"💰 Precio: ${price:.4f}\n"
            f"📉 RSI: {current['rsi']:.2f}\n"
            f"⏰ Hora: {self.df.index[idx].strftime('%Y-%m-%d %H:%M:%S')}\n"
        )
        await self.send_telegram_message(message)

    async def notify_entry(self, position: Position) -> None:
        """Notificar apertura de posición."""
        pnl_target = position.margin_used * (self.take_profit_percent_long if position.direction == 'LONG' else self.take_profit_percent_short) * self.leverage
        
        message = (
            f"✅ <b>POSICIÓN ABIERTA</b>\n\n"
            f"📊 Par: {self.symbol}\n"
            f"📈 Dirección: <b>{position.direction}</b>\n"
            f"💰 Precio entrada: ${position.entry_price:.4f}\n"
            f"📏 Tamaño: {position.size:.4f}\n"
            f"💵 Margen usado: ${position.margin_used:.2f}\n"
            f"🎯 Take Profit: ${position.take_profit:.4f} (+${pnl_target:.2f})\n"
            f"🛑 Stop Loss: ${position.stop_loss:.4f}\n"
            f"⚠️ Liquidación: ${position.liquidation_price:.4f}\n"
            f"⏰ Hora: {position.entry_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        )
        await self.send_telegram_message(message)

    async def notify_exit(self, trade: Dict) -> None:
        """Notificar cierre de posición."""
        emoji = "🟢" if trade['pnl'] > 0 else "🔴"
        
        message = (
            f"{emoji} <b>POSICIÓN CERRADA</b>\n\n"
            f"📊 Par: {self.symbol}\n"
            f"📈 Dirección: {trade['direction']}\n"
            f"💰 Entrada: ${trade['entry_price']:.4f}\n"
            f"💰 Salida: ${trade['exit_price']:.4f}\n"
            f"💵 P/L: ${trade['pnl']:.2f} ({trade['pnl_pct']:.2f}%)\n"
            f"📝 Razón: {trade['exit_reason']}\n"
            f"💼 Balance: ${self.balance:.2f}\n"
            f"⏰ Duración: {trade['duration_candles']} velas\n"
        )
        await self.send_telegram_message(message)

    async def notify_daily_summary(self) -> None:
        """Enviar resumen diario."""
        if not self.trades:
            return
        
        today_trades = [t for t in self.trades if t['entry_time'].date() == datetime.now().date()]
        if not today_trades:
            return
        
        winning = sum(1 for t in today_trades if t['pnl'] > 0)
        losing = len(today_trades) - winning
        total_pnl = sum(t['pnl'] for t in today_trades)
        win_rate = (winning / len(today_trades) * 100) if today_trades else 0
        
        message = (
            f"📊 <b>RESUMEN DEL DÍA</b>\n\n"
            f"📈 Total trades: {len(today_trades)}\n"
            f"✅ Ganadores: {winning}\n"
            f"❌ Perdedores: {losing}\n"
            f"🎯 Win Rate: {win_rate:.1f}%\n"
            f"💵 P/L Total: ${total_pnl:.2f}\n"
            f"💼 Balance: ${self.balance:.2f}\n"
            f"📈 Retorno: {((self.balance - self.initial_balance) / self.initial_balance * 100):.2f}%\n"
        )
        await self.send_telegram_message(message)

    # ----------------------------------------------------------------------------------
    # Descarga y preparación de datos
    # ----------------------------------------------------------------------------------
    async def fetch_candles(self) -> Optional[pd.DataFrame]:
        """Descargar velas recientes desde Binance Futures."""

        try:
            ohlcv = self.exchange.fetch_ohlcv(
                symbol=self.symbol,
                timeframe=self.timeframe,
                limit=self.candle_limit
            )
        except Exception as exc:
            logger.error(f"Error al descargar velas: {exc}")
            return None

        df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms', utc=True).dt.tz_convert('America/Argentina/Buenos_Aires')
        df.set_index('timestamp', inplace=True)

        return df

    def calculate_indicators(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calcular los mismos indicadores que en el backtest."""

        df = df.copy()

        df['ema_9'] = df['close'].ewm(span=9, adjust=False).mean()
        df['ema_21'] = df['close'].ewm(span=21, adjust=False).mean()
        df['ema_50'] = df['close'].ewm(span=50, adjust=False).mean()
        df['ema_200'] = df['close'].ewm(span=200, adjust=False).mean()

        delta = df['close'].diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
        rs = gain / loss
        df['rsi'] = 100 - (100 / (1 + rs))

        exp1 = df['close'].ewm(span=12, adjust=False).mean()
        exp2 = df['close'].ewm(span=26, adjust=False).mean()
        df['macd'] = exp1 - exp2
        df['macd_signal'] = df['macd'].ewm(span=9, adjust=False).mean()
        df['macd_hist'] = df['macd'] - df['macd_signal']

        high_low = df['high'] - df['low']
        high_close = np.abs(df['high'] - df['close'].shift())
        low_close = np.abs(df['low'] - df['close'].shift())
        ranges = pd.concat([high_low, high_close, low_close], axis=1)
        true_range = np.max(ranges, axis=1)
        df['atr'] = true_range.rolling(14).mean()

        df['bb_middle'] = df['close'].rolling(window=20).mean()
        df['bb_std'] = df['close'].rolling(window=20).std()
        df['bb_upper'] = df['bb_middle'] + (df['bb_std'] * 2)
        df['bb_lower'] = df['bb_middle'] - (df['bb_std'] * 2)

        return df

    async def update_market_data(self) -> bool:
        """Actualizar `self.df` con las últimas velas e indicadores."""

        df = await self.fetch_candles()
        if df is None or df.empty:
            return False

        df = self.calculate_indicators(df)
        self.df = df
        return True

    # ----------------------------------------------------------------------------------
    # Detección de momentum fuerte
    # ----------------------------------------------------------------------------------
    def detect_strong_momentum(self, idx: int) -> Optional[str]:
        """Detecta caídas o subidas fuertes con volumen para entrar DURANTE el movimiento."""
        if idx < 50:
            return None
        
        current = self.df.iloc[idx]
        recent = self.df.iloc[idx - self.momentum_candles_window:idx + 1]
        volume_window = self.df.iloc[max(0, idx - 20):idx]
        
        # Calcular cambio de precio en la ventana
        price_change_pct = ((current['close'] - recent['close'].iloc[0]) / recent['close'].iloc[0]) * 100
        
        # Volumen promedio
        avg_volume = volume_window['volume'].mean()
        current_volume = current['volume']
        
        # Verificar volumen fuerte
        strong_volume = current_volume >= avg_volume * self.momentum_volume_multiplier
        
        # CAÍDA FUERTE (entrada LONG siguiendo la caída)
        if price_change_pct <= -self.momentum_threshold_pct and strong_volume:
            # Verificar que está cayendo activamente
            bearish_candles = sum(1 for i in range(len(recent)) if recent['close'].iloc[i] < recent['open'].iloc[i])
            if bearish_candles >= 2:
                # NUEVA LÓGICA: Entrar DURANTE la caída, no esperar reversión
                # Solo verificar que no está en rebote fuerte
                if current['rsi'] < 45:  # Sobrevendido o neutral bajo
                    # Verificar que la caída continúa (vela actual también bajista o neutral)
                    if current['close'] <= recent['close'].iloc[-2]:  # Precio sigue bajando
                        logger.info(f"🔥 MOMENTUM BAJISTA DETECTADO: Caída {price_change_pct:.2f}% - Entrada LONG siguiendo caída")
                        return 'LONG_MOMENTUM'
        
        # SUBIDA FUERTE (entrada SHORT siguiendo la subida)
        elif price_change_pct >= self.momentum_threshold_pct and strong_volume:
            # Verificar que está subiendo activamente
            bullish_candles = sum(1 for i in range(len(recent)) if recent['close'].iloc[i] > recent['open'].iloc[i])
            if bullish_candles >= 2:
                # NUEVA LÓGICA: Entrar DURANTE la subida, no esperar reversión
                if current['rsi'] > 55:  # Sobrecomprado o neutral alto
                    # Verificar que la subida continúa
                    if current['close'] >= recent['close'].iloc[-2]:  # Precio sigue subiendo
                        logger.info(f"🔥 MOMENTUM ALCISTA DETECTADO: Subida {price_change_pct:.2f}% - Entrada SHORT siguiendo subida")
                        return 'SHORT_MOMENTUM'
        
        return None
    
    def detect_momentum_reversal(self, idx: int) -> bool:
        """Detecta si el momentum se está revirtiendo para cerrar la posición."""
        if not self.position or self.position.entry_type != 'momentum':
            return False
        
        current = self.df.iloc[idx]
        prev = self.df.iloc[idx - 1] if idx > 0 else current
        
        if self.position.direction == 'LONG':
            # Cerrar LONG si detecta reversión bajista
            reversal_signals = [
                current['close'] < current['ema_9'],  # Precio bajo EMA rápida
                current['ema_9'] < current['ema_21'],  # Cruce bajista
                current['macd_hist'] < prev['macd_hist'],  # MACD perdiendo fuerza
                current['close'] < current['open'],  # Vela bajista
                current['rsi'] > 65  # Sobrecomprado
            ]
            return sum(reversal_signals) >= 3
        
        else:  # SHORT
            # Cerrar SHORT si detecta reversión alcista
            reversal_signals = [
                current['close'] > current['ema_9'],  # Precio sobre EMA rápida
                current['ema_9'] > current['ema_21'],  # Cruce alcista
                current['macd_hist'] > prev['macd_hist'],  # MACD ganando fuerza
                current['close'] > current['open'],  # Vela alcista
                current['rsi'] < 35  # Sobrevendido
            ]
            return sum(reversal_signals) >= 3
    
    # ----------------------------------------------------------------------------------
    # Reglas de señal idénticas al backtest
    # ----------------------------------------------------------------------------------
    def is_long_trading_hours(self, timestamp: pd.Timestamp) -> bool:
        if timestamp.weekday() >= 5:
            return False

        current_time = timestamp.time()
        start_time = time(8, 0)
        end_time = time(20, 0)
        return start_time <= current_time <= end_time

    def detect_mean_reversion_long(self, idx: int) -> bool:
        if idx < 200:
            return False

        current = self.df.iloc[idx]
        prev = self.df.iloc[idx - 1]
        recent = self.df.iloc[max(0, idx - 20):idx]

        if current['ema_50'] <= current['ema_200']:
            return False
        if len(recent) >= 2 and recent['ema_200'].iloc[-1] <= recent['ema_200'].iloc[0]:
            return False

        at_support = current['close'] < current['bb_lower'] * 1.003
        if not at_support:
            return False

        if not (current['close'] > current['open'] and current['close'] >= prev['close']):
            return False

        volume_mean = recent['volume'].mean() if len(recent) > 0 else current['volume']

        oversold_conditions = [
            current['ema_9'] < current['ema_21'],
            current['close'] < current['ema_21'],
            current['rsi'] < 35,
            current['macd'] < current['macd_signal'],
            current['macd_hist'] > prev['macd_hist'],
            current['volume'] >= volume_mean * 1.1,
            current['close'] >= current['ema_200'] * 0.99
        ]

        return sum(oversold_conditions) >= 5

    def detect_mean_reversion_short_improved(self, idx: int) -> bool:
        if idx < 200:
            return False

        current = self.df.iloc[idx]
        prev = self.df.iloc[idx - 1]
        recent = self.df.iloc[max(0, idx - 20):idx]
        recent_longer = self.df.iloc[max(0, idx - 50):idx]

        uptrend_context = [
            current['ema_9'] > current['ema_21'],
            current['close'] > current['ema_50'],
            recent_longer['close'].iloc[-1] > recent_longer['close'].iloc[0]
        ]

        if sum(uptrend_context) < 1:
            self.filtered_short_signals += 1
            return False

        bb_distance = (current['close'] - current['bb_upper']) / current['bb_upper'] * 100
        near_bb_upper = bb_distance > -1.0
        if not near_bb_upper:
            return False

        candle_range = current['high'] - current['low']
        if candle_range == 0:
            return False

        close_position = (current['close'] - current['low']) / candle_range
        has_bearish_pattern = close_position < 0.7
        if not has_bearish_pattern:
            return False

        volume_mean = recent['volume'].mean()
        short_conditions = [
            current['rsi'] > 65,
            bb_distance > -0.8,
            current['close'] > current['ema_9'],
            current['macd'] > current['macd_signal'],
            current['macd_hist'] < prev['macd_hist'],
            current['volume'] > volume_mean * 1.05,
            current['high'] >= recent['high'].tail(10).max(),
            (current['close'] - current['ema_50']) / current['ema_50'] * 100 > 0.5
        ]

        return sum(short_conditions) >= 4

    def add_short_quality_filter(self, idx: int) -> bool:
        if idx < 10:
            return True

        recent = self.df.iloc[max(0, idx - 10):idx]

        for i in range(max(0, len(recent) - 3), len(recent)):
            if i < len(recent) - 1:
                if recent['ema_9'].iloc[i] < recent['ema_21'].iloc[i] and \
                   recent['ema_9'].iloc[i + 1] >= recent['ema_21'].iloc[i + 1]:
                    self.filtered_short_signals += 1
                    return False

        volumes = recent['volume'].tail(3).values
        if len(volumes) >= 3:
            if volumes[-1] > volumes[-2] * 2.0 and volumes[-2] > volumes[-3] * 2.0:
                self.filtered_short_signals += 1
                return False

        return True

    def analyze_market_direction(self, idx: int) -> Optional[str]:
        if idx < 200:
            return None

        current_time = self.df.index[idx]
        
        # PRIORIDAD 1: Detectar momentum fuerte (sin restricción horaria)
        if self.momentum_enabled:
            momentum_signal = self.detect_strong_momentum(idx)
            if momentum_signal:
                return momentum_signal

        # PRIORIDAD 2: Mean reversion tradicional
        if self.detect_mean_reversion_long(idx):
            if self.is_long_trading_hours(current_time):
                return 'LONG'
            self.filtered_long_signals += 1
            return None

        if self.detect_mean_reversion_short_improved(idx) and self.add_short_quality_filter(idx):
            return 'SHORT'

        return None

    # ----------------------------------------------------------------------------------
    # Gestión de órdenes y posiciones
    # ----------------------------------------------------------------------------------
    def calculate_stop_loss(self, entry_price: float, direction: str) -> float:
        if direction == 'LONG':
            return entry_price - entry_price * self.stop_loss_percent_long
        return entry_price + entry_price * self.stop_loss_percent_short

    def calculate_liquidation_price(self, entry_price: float, direction: str) -> float:
        liquidation_distance = entry_price * (1 / self.leverage)
        if direction == 'LONG':
            return entry_price - liquidation_distance
        return entry_price + liquidation_distance

    def calculate_take_profit(
        self,
        entry_price: float,
        direction: str,
        is_first_trade: bool,
        idx: int
    ) -> float:
        if direction == 'LONG':
            return entry_price + entry_price * self.take_profit_percent_long
        return entry_price - entry_price * self.take_profit_percent_short

    def calculate_position_size(self, entry_price: float) -> tuple[float, float]:
        margin = self.balance * self.risk_per_trade
        position_value = margin * self.leverage
        size = position_value / entry_price if entry_price > 0 else 0
        return size, margin

    async def open_position(self, direction: str, idx: int) -> None:
        """Registrar señal sin abrir órdenes reales."""

        current_candle = self.df.iloc[idx]
        entry_price = current_candle['close']
        
        # Determinar tipo de entrada
        is_momentum = direction.endswith('_MOMENTUM')
        if is_momentum:
            direction = direction.replace('_MOMENTUM', '')  # Limpiar dirección
            entry_type = 'momentum'
        else:
            entry_type = 'mean_reversion'

        size, margin = self.calculate_position_size(entry_price)
        if size <= 0:
            logger.error("No se pudo calcular un tamaño de posición válido")
            return

        # Stop loss MÁS AMPLIO para momentum (seguir el movimiento completo)
        if entry_type == 'momentum':
            if direction == 'LONG':
                stop_loss = entry_price - entry_price * 0.025  # 2.5% SL (más amplio)
                take_profit = entry_price + entry_price * 0.05  # 5% TP (más ambicioso)
            else:
                stop_loss = entry_price + entry_price * 0.028  # 2.8% SL (más amplio)
                take_profit = entry_price - entry_price * 0.045  # 4.5% TP (más ambicioso)
        else:
            stop_loss = self.calculate_stop_loss(entry_price, direction)
            take_profit = self.calculate_take_profit(
                entry_price,
                direction,
                is_first_trade=self.daily_first_trade,
                idx=idx
            )
        
        liquidation_price = self.calculate_liquidation_price(entry_price, direction)

        entry_time = self.df.index[idx]

        self.position = Position(
            direction=direction,
            size=size,
            entry_price=float(entry_price),
            entry_time=entry_time.to_pydatetime(),
            entry_idx=idx,
            stop_loss=stop_loss,
            take_profit=take_profit,
            liquidation_price=liquidation_price,
            is_first_trade=self.daily_first_trade,
            margin_used=margin,
            entry_type=entry_type
        )

        logger.info(
            f"📢 Señal {direction} ({entry_type.upper()}): tamaño simulado {size:.4f} @ {entry_price:.4f} | "
            f"SL {stop_loss:.4f} | TP {take_profit:.4f} | Liq {liquidation_price:.4f}"
        )
        
        # Enviar notificación de entrada
        await self.notify_entry(self.position)

        self.last_signal_time = entry_time

        current_date = entry_time.date()
        if self.last_trade_date != current_date:
            self.daily_first_trade = False if self.daily_first_trade else self.daily_first_trade
            self.last_trade_date = current_date
        else:
            if self.daily_first_trade and entry_time.time() >= time(12, 0):
                self.daily_first_trade = False

    async def close_position(self, exit_price: float, reason: str) -> None:
        if not self.position:
            return

        position = self.position

        if position.direction == 'LONG':
            pnl = (exit_price - position.entry_price) * position.size
        else:
            pnl = (position.entry_price - exit_price) * position.size

        pnl_pct = (pnl / (position.entry_price * position.size / self.leverage)) * 100

        self.balance += pnl
        
        trade = {
            'entry_time': position.entry_time,
            'exit_time': datetime.now(),
            'direction': position.direction,
            'entry_price': position.entry_price,
            'exit_price': exit_price,
            'stop_loss': position.stop_loss,
            'take_profit': position.take_profit,
            'liquidation_price': position.liquidation_price,
            'position_size': position.size,
            'margin_used': position.margin_used,
            'pnl': pnl,
            'pnl_pct': pnl_pct,
            'exit_reason': reason,
            'duration_candles': len(self.df) - 1 - position.entry_idx,
            'entry_type': position.entry_type
        }
        
        self.trades.append(trade)

        self.equity_curve.append({
            'time': datetime.now(),
            'balance': self.balance
        })

        logger.info(
            f"ℹ️ Señal cerrada ({reason}). PnL simulado: {pnl:.2f} USDT | PnL%: {pnl_pct:.2f}% | Balance simulado: {self.balance:.2f}"
        )
        
        # Guardar en Excel
        self._save_trade_to_excel(trade)
        
        # Enviar notificación de salida
        await self.notify_exit(trade)

        self.position = None

    async def manage_position(self, idx: int) -> None:
        if not self.position:
            return

        current_candle = self.df.iloc[idx]
        candles_elapsed = idx - self.position.entry_idx
        
        # Gestión especial para posiciones de momentum
        if self.position.entry_type == 'momentum':
            # Actualizar trailing stop
            current_price = current_candle['close']
            
            if self.position.direction == 'LONG':
                # Calcular ganancia actual
                profit_pct = ((current_price - self.position.entry_price) / self.position.entry_price) * 100
                
                # Activar trailing stop si hay ganancia suficiente
                if profit_pct >= self.trailing_stop_activation_pct:
                    # Calcular nuevo trailing stop
                    max_price = max(current_candle['high'], self.position.entry_price * (1 + profit_pct / 100))
                    new_trailing = max_price * (1 - self.trailing_stop_distance_pct / 100)
                    
                    # Actualizar trailing stop (solo sube, nunca baja)
                    if self.position.trailing_stop is None or new_trailing > self.position.trailing_stop:
                        self.position.trailing_stop = new_trailing
                        logger.info(f"📈 Trailing stop actualizado: ${new_trailing:.4f} (ganancia: {profit_pct:.2f}%)")
                
                # Verificar trailing stop
                if self.position.trailing_stop and current_candle['low'] <= self.position.trailing_stop:
                    await self.close_position(self.position.trailing_stop, 'Trailing Stop (momentum)')
                    return
                
                # Verificar reversión de momentum
                if self.detect_momentum_reversal(idx):
                    await self.close_position(current_price, 'Reversión de Momentum')
                    return
            
            else:  # SHORT
                # Calcular ganancia actual
                profit_pct = ((self.position.entry_price - current_price) / self.position.entry_price) * 100
                
                # Activar trailing stop si hay ganancia suficiente
                if profit_pct >= self.trailing_stop_activation_pct:
                    # Calcular nuevo trailing stop
                    min_price = min(current_candle['low'], self.position.entry_price * (1 - profit_pct / 100))
                    new_trailing = min_price * (1 + self.trailing_stop_distance_pct / 100)
                    
                    # Actualizar trailing stop (solo baja, nunca sube)
                    if self.position.trailing_stop is None or new_trailing < self.position.trailing_stop:
                        self.position.trailing_stop = new_trailing
                        logger.info(f"📉 Trailing stop actualizado: ${new_trailing:.4f} (ganancia: {profit_pct:.2f}%)")
                
                # Verificar trailing stop
                if self.position.trailing_stop and current_candle['high'] >= self.position.trailing_stop:
                    await self.close_position(self.position.trailing_stop, 'Trailing Stop (momentum)')
                    return
                
                # Verificar reversión de momentum
                if self.detect_momentum_reversal(idx):
                    await self.close_position(current_price, 'Reversión de Momentum')
                    return
        
        # Gestión estándar para todas las posiciones
        if self.position.direction == 'LONG':
            if current_candle['low'] <= self.position.liquidation_price:
                await self.close_position(self.position.liquidation_price, 'LIQUIDACIÓN (simulada)')
                return
            if current_candle['low'] <= self.position.stop_loss:
                await self.close_position(self.position.stop_loss, 'Stop Loss (simulado)')
                return
            if current_candle['high'] >= self.position.take_profit:
                await self.close_position(self.position.take_profit, 'Take Profit (simulado)')
                return

        else:
            if current_candle['high'] >= self.position.liquidation_price:
                await self.close_position(self.position.liquidation_price, 'LIQUIDACIÓN (simulada)')
                return
            if current_candle['high'] >= self.position.stop_loss:
                await self.close_position(self.position.stop_loss, 'Stop Loss (simulado)')
                return
            if current_candle['low'] <= self.position.take_profit:
                await self.close_position(self.position.take_profit, 'Take Profit (simulado)')
                return

    # ----------------------------------------------------------------------------------
    # Bucle principal
    # ----------------------------------------------------------------------------------
    async def run(self) -> None:
        """Ejecutar el ciclo principal del bot."""

        self.is_running = True
        logger.info("=== Iniciando Hybrid Futures Trading Bot (15x) ===")
        logger.info(f"Balance inicial simulado: {self.initial_balance} USDT | Riesgo por trade: {self.risk_per_trade*100:.1f}%")

        while self.is_running:
            has_data = await self.update_market_data()
            if not has_data:
                logger.warning("No se pudieron actualizar las velas. Reintentando...")
                await asyncio.sleep(self.refresh_seconds)
                continue

            if self.df is None or len(self.df) < 200:
                logger.info("Esperando suficientes velas para evaluar señales (>=200)...")
                await asyncio.sleep(self.refresh_seconds)
                continue

            idx = len(self.df) - 1
            await self.manage_position(idx)

            if self.position is None:
                current_time = self.df.index[idx]
                current_date = current_time.date()
                if self.last_trade_date != current_date:
                    self.daily_first_trade = True
                    self.last_trade_date = current_date

                direction = self.analyze_market_direction(idx)

                if direction and (self.last_signal_time is None or current_time > self.last_signal_time):
                    await self.open_position(direction, idx)

            await asyncio.sleep(self.refresh_seconds)

        logger.info("Bot detenido correctamente.")

    def stop(self) -> None:
        self.is_running = False


async def main() -> None:
    """Punto de entrada. Lee `config.json` y lanza el bot."""

    try:
        with open('config.json', 'r', encoding='utf-8') as cfg_file:
            config = json.load(cfg_file)
    except FileNotFoundError:
        logger.error(
            "Archivo config.json no encontrado. Crea uno con: {\"api_key\":..., \"api_secret\":..., \"telegram_token\":..., \"telegram_chat_id\":..., \"symbol\":\"HYPE/USDT\", \"initial_balance\":1000}"
        )
        return

    api_key = config.get('api_key')
    api_secret = config.get('api_secret')
    telegram_token = config.get('telegram_token')
    telegram_chat_id = config.get('telegram_chat_id')
    
    if not api_key or not api_secret:
        logger.error("Config inválido: faltan `api_key` o `api_secret`.")
        return
    
    if not telegram_token or not telegram_chat_id:
        logger.error("Config inválido: faltan `telegram_token` o `telegram_chat_id`.")
        return

    symbol = config.get('symbol', 'HYPE/USDT')
    initial_balance = float(config.get('initial_balance', 1_000))
    leverage = int(config.get('leverage', 15))
    risk_per_trade = float(config.get('risk_per_trade', 0.10))
    refresh_seconds = int(config.get('refresh_seconds', 60))

    bot = HybridFuturesTradingBot(
        api_key=api_key,
        api_secret=api_secret,
        telegram_token=telegram_token,
        telegram_chat_id=telegram_chat_id,
        symbol=symbol,
        leverage=leverage,
        risk_per_trade=risk_per_trade,
        initial_balance=initial_balance,
        refresh_seconds=refresh_seconds
    )

    try:
        await bot.run()
    except KeyboardInterrupt:
        logger.info("Interrupción manual recibida. Cerrando bot...")
        bot.stop()


if __name__ == "__main__":
    asyncio.run(main())
